"""
Holdings-file upload pipeline.

The upload flow is two-step:

1. **Preview** (``upload_preview``). User picks a CSV or XLSX. We parse
   it just enough to introspect: detect encoding (CSV), detect delimiter
   (CSV), enumerate sheets (XLSX), and return the first ~200 rows as a
   2-D string grid. We stash the parsed-but-unmapped grid under a UUID
   token in :data:`~porxpy.config.UPLOAD_DIR` so step 2 doesn't have to
   re-parse the upload.

2. **Commit** (``upload_commit``). User has chosen the sheet, the header
   row, the weight unit, the decimal notation, and a column→canonical-
   field mapping. We re-read the token, apply the mapping, normalise
   each row to the unified holdings superset schema (stamping a stable
   ``_row_id`` on each), and write the result into the fund's single
   ``holdings`` cache slot — with ``source="manual_upload"`` so the
   look-through cards / rollup / holdings editor can tell where the
   data came from. The Yahoo top-10 fetch path never overwrites a
   ``manual_upload`` blob.

The server-side temp store is reaped on every preview call (any tokens
older than :data:`~porxpy.config.UPLOAD_TOKEN_TTL_MIN` get deleted) and
also explicitly after a successful commit.
"""

from __future__ import annotations

import csv
import io
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from porxpy.config import (
    CACHE_CATEGORIES,
    DEFAULT_CACHE_CONFIG,
    ENRICHABLE_FIELDS,
    UPLOAD_DIR,
    UPLOAD_TOKEN_TTL_MIN,
)
from porxpy.resolver import clean_holding_ticker_input
from porxpy.resources import (
    country_to_currency,
    country_to_mstar,
    resolve_currency,
    resolve_sector,
    resolve_sub_class,
)
from porxpy.utils import (
    age_days, cache_put, cache_read, cache_write, coerce_holdings_row,
    default_holding_asset_class, normalize_holding_asset_class, now_iso,
)


# ---------------------------------------------------------------------------
# Token store — UUID → JSON file under UPLOAD_DIR
# ---------------------------------------------------------------------------
UPLOAD_DIR.mkdir(exist_ok=True)


# Maximum number of rows we'll process for Yahoo enrichment in a single
# commit. Above this, enrichment runs on the first N rows; the rest get
# default-fill only. Bounded so a 5000-row holdings file doesn't park
# the request thread for minutes on a cold cache.
_ENRICH_ROW_CAP = 1000

# Fields that may appear in the ``enrich_fields`` list on a commit.
# Mirrors the canonical :data:`~porxpy.config.ENRICHABLE_FIELDS` — the
# upload dialog and the post-upload "Enrich through Yahoo" button share
# one vocabulary so the user's options match across the two paths.
_ENRICHABLE_FIELDS = set(ENRICHABLE_FIELDS)


# ---------------------------------------------------------------------------
# Cancel registry (v0.15.4)
# ---------------------------------------------------------------------------
# Some commits — specifically Yahoo enrichment on a cold cache — take
# minutes. The user wants a Cancel button on the upload dialog that
# actually stops the work. We can't interrupt a running Python loop
# from a different thread without cooperation, so the loop polls a
# token-keyed registry instead.
#
# Flow:
#   1. /api/upload/cancel  POST {token} → mark_cancelled(token).
#   2. The active upload_commit's enrichment loop calls is_cancelled
#      at the top of each iteration; on True it raises UploadCancelled
#      BEFORE any cache write, so the on-disk holdings stay untouched.
#   3. upload_commit always clears the flag on exit (success, normal
#      error, or cancel) so a retry with the same token starts clean.
#
# Per-symbol Yahoo calls already-completed during the cancelled commit
# remain in the symbol-info cache — that's a feature, not a leak.
import threading

_CANCELLED_TOKENS: set[str] = set()
_CANCEL_LOCK = threading.Lock()


class UploadCancelled(Exception):
    """Raised by upload_commit when its token was marked for cancel.

    Caught by the /api/upload/commit route, which returns a 200 with
    ``{"cancelled": true}`` so the frontend can close the dialog
    silently rather than show an error.
    """
    def __init__(self, token: str):
        super().__init__(f"upload cancelled: {token}")
        self.token = token


def mark_cancelled(token: str) -> None:
    """Flag ``token`` so the active upload_commit raises at its next check."""
    if not token:
        return
    with _CANCEL_LOCK:
        _CANCELLED_TOKENS.add(token)


def is_cancelled(token: str) -> bool:
    """True iff ``token`` is in the cancel registry."""
    if not token:
        return False
    with _CANCEL_LOCK:
        return token in _CANCELLED_TOKENS


def clear_cancel(token: str) -> None:
    """Remove ``token`` from the registry (used on commit exit)."""
    if not token:
        return
    with _CANCEL_LOCK:
        _CANCELLED_TOKENS.discard(token)


def _token_path(token: str) -> Path:
    """Return the absolute path of the temp file for ``token``."""
    # Defensive: only allow plain-ish UUID-shaped strings as filenames.
    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", token)
    return UPLOAD_DIR / f"{safe}.json"


def _reap_expired() -> int:
    """Delete temp files older than :data:`UPLOAD_TOKEN_TTL_MIN` minutes.

    Called at the start of every preview to keep ``UPLOAD_DIR`` from
    growing without bound. Returns the number of files removed.
    """
    cutoff_minutes = UPLOAD_TOKEN_TTL_MIN
    removed = 0
    now_ts = datetime.now(timezone.utc)
    for fp in UPLOAD_DIR.glob("*.json"):
        try:
            mtime = datetime.fromtimestamp(fp.stat().st_mtime, tz=timezone.utc)
            if (now_ts - mtime).total_seconds() / 60 > cutoff_minutes:
                fp.unlink()
                removed += 1
        except Exception:
            pass
    return removed


def _save_token(token: str, payload: dict) -> None:
    """Persist a preview payload under ``token``."""
    fp = _token_path(token)
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)


def _load_token(token: str) -> dict | None:
    """Read a preview payload by ``token``.

    Returns ``None`` if the token is missing, malformed, or older than
    :data:`UPLOAD_TOKEN_TTL_MIN`.
    """
    fp = _token_path(token)
    if not fp.exists():
        return None
    try:
        # TTL check via file mtime — independent of the file's own
        # ``created_at`` field, in case the clock skewed.
        mtime = datetime.fromtimestamp(fp.stat().st_mtime, tz=timezone.utc)
        if (datetime.now(timezone.utc) - mtime).total_seconds() / 60 > UPLOAD_TOKEN_TTL_MIN:
            try:
                fp.unlink()
            except Exception:
                pass
            return None
        with open(fp, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _delete_token(token: str) -> None:
    """Remove a preview token file (no-op if already gone)."""
    fp = _token_path(token)
    try:
        if fp.exists():
            fp.unlink()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# CSV parsing — encoding + delimiter detection
# ---------------------------------------------------------------------------
_ENCODINGS_TO_TRY: tuple[str, ...] = ("utf-8", "utf-8-sig", "cp1252", "latin-1")


def _decode_csv_bytes(data: bytes) -> tuple[str, str]:
    """Decode CSV bytes by trying a small ladder of encodings.

    Args:
        data: Raw bytes from the upload.

    Returns:
        ``(decoded_text, encoding_used)``. ``latin-1`` always succeeds
        (it accepts any byte) so this never raises.
    """
    for enc in _ENCODINGS_TO_TRY:
        try:
            return data.decode(enc), enc
        except UnicodeDecodeError:
            continue
    # Should be unreachable since latin-1 accepts anything.
    return data.decode("latin-1", errors="replace"), "latin-1"


def _sniff_delimiter(sample: str) -> str:
    """Guess the CSV delimiter from a sample of the file.

    Approach: csv.Sniffer first, but it fails on files with preamble
    (the early lines have no delimiter, which throws it off). When
    Sniffer can't decide, fall back to a tally — count the most
    consistent delimiter across the bulk of the lines, ignoring lines
    where any candidate appears 0 times.

    Most issuer files use one of ``,`` ``;`` ``\\t`` ``|``; default to
    comma if nothing produces a clear winner.
    """
    candidates = ",;\t|"
    try:
        dialect = csv.Sniffer().sniff(sample[:8192], delimiters=candidates)
        return dialect.delimiter
    except csv.Error:
        pass

    # Tally fallback. For each candidate, look at lines where it appears
    # at least once, and find the most-common count among those lines.
    # The candidate with the highest "best mode count" wins. This handles
    # preamble cleanly because preamble lines have count 0 for every
    # candidate and don't contribute to any tally.
    lines = [ln for ln in sample.split("\n") if ln.strip()][:200]
    best_delim = ","
    best_score = 0
    for delim in candidates:
        from collections import Counter
        counts = [ln.count(delim) for ln in lines if ln.count(delim) > 0]
        if not counts:
            continue
        mode_count, mode_freq = Counter(counts).most_common(1)[0]
        # Score: how many lines hit the mode count; require ≥ 3 lines and
        # the mode count itself ≥ 1.
        if mode_freq >= 3 and mode_count >= 1 and mode_freq > best_score:
            best_score = mode_freq
            best_delim = delim
    return best_delim


def _parse_csv(data: bytes, delimiter: str | None = None
               ) -> tuple[list[list[str]], dict]:
    """Parse a CSV upload into a 2-D string grid.

    Args:
        data: Raw upload bytes.
        delimiter: Optional explicit delimiter override. ``None`` → sniff.

    Returns:
        ``(rows, info)`` where ``rows`` is a list of row-lists (strings,
        no trimming), and ``info`` is ``{encoding, delimiter, sheets:None}``.
    """
    text, encoding = _decode_csv_bytes(data)
    delim = delimiter if (delimiter and delimiter in ",;\t|") else _sniff_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = [list(r) for r in reader]
    return rows, {
        "encoding":  encoding,
        "delimiter": delim,
        "sheets":    None,
    }


# ---------------------------------------------------------------------------
# XLSX parsing — sheet enumeration + cell extraction
# ---------------------------------------------------------------------------
def _parse_xlsx(data: bytes, sheet_name: str | None = None
                ) -> tuple[list[list[str]], dict]:
    """Parse an XLSX upload into a 2-D string grid.

    Args:
        data: Raw upload bytes.
        sheet_name: Optional sheet to read. If absent, picks the first
            sheet with at least 5 rows of any data — Amundi/Vanguard
            style "Disclaimer" sheets at index 0 with a few cells of
            legal text get skipped, and Holdings ends up selected.

    Returns:
        ``(rows, info)`` where ``info`` carries ``{encoding: None,
        delimiter: None, sheets, picked_sheet}``.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False,
    )
    sheets = wb.sheetnames
    if not sheets:
        raise ValueError("workbook has no sheets")

    if sheet_name and sheet_name in sheets:
        picked = sheet_name
    else:
        # Auto-pick: first sheet with ≥5 non-empty rows.
        picked = None
        for sn in sheets:
            ws = wb[sn]
            non_empty = 0
            for row in ws.iter_rows(values_only=True):
                if any(c not in (None, "") for c in row):
                    non_empty += 1
                    if non_empty >= 5:
                        break
            if non_empty >= 5:
                picked = sn
                break
        if picked is None:
            picked = sheets[0]

    ws = wb[picked]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])

    return rows, {
        "encoding":     None,
        "delimiter":    None,
        "sheets":       sheets,
        "picked_sheet": picked,
    }


# ---------------------------------------------------------------------------
# Header-row autodetection
# ---------------------------------------------------------------------------
def _detect_header_row(rows: list[list[str]]) -> int:
    """Guess which row is the header.

    Heuristic: the first row where (a) at least 3 cells are non-empty,
    and (b) no cell in that row parses as a pure number. Issuer files
    typically have a few rows of preamble (fund name, "as of <date>",
    legal blurb) before the actual table starts.

    Args:
        rows: 2-D string grid as returned by parse_csv / parse_xlsx.

    Returns:
        Zero-based row index. Defaults to ``0`` if nothing matches.
    """
    for i, row in enumerate(rows[:30]):  # never look beyond the first 30 rows
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) < 3:
            continue
        # If every non-empty cell is a number, it's data, not headers.
        all_numeric = True
        for c in non_empty:
            try:
                float(c.replace(",", "").replace("%", "").strip())
            except (TypeError, ValueError):
                all_numeric = False
                break
        if not all_numeric:
            return i
    return 0


# ---------------------------------------------------------------------------
# Number parsing — decimal notation + weight unit
# ---------------------------------------------------------------------------
def _parse_number(raw: str, decimal: str = "auto") -> float | None:
    """Parse a possibly-quoted, possibly-formatted number string.

    Supports:

    * ``decimal="dot"`` — comma is a thousands separator, dot is decimal.
    * ``decimal="comma"`` — dot is a thousands separator, comma is decimal.
    * ``decimal="auto"`` — guess per-string: if the rightmost separator
      has 1-3 digits after it, that's the decimal point.

    Strips whitespace, ``%``, currency symbols, and parentheses (negative
    accountancy notation). Returns ``None`` for unparseable input.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    # Strip percent / common currency prefixes / accountancy parens
    s = s.replace("%", "")
    for sym in ("€", "$", "£", "¥", "kr"):
        s = s.replace(sym, "")
    is_neg = False
    if s.startswith("(") and s.endswith(")"):
        is_neg = True
        s = s[1:-1]
    s = s.strip()
    if not s:
        return None

    # Apply decimal notation
    if decimal == "dot":
        s = s.replace(",", "")  # thousands separator only
    elif decimal == "comma":
        s = s.replace(".", "").replace(",", ".")
    else:  # auto
        # If the string has both, the rightmost is decimal.
        last_dot   = s.rfind(".")
        last_comma = s.rfind(",")
        if last_dot >= 0 and last_comma >= 0:
            if last_comma > last_dot:
                # Comma is rightmost → decimal=comma
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif last_comma >= 0:
            # Only comma. If it has exactly 3 digits after it, treat as
            # thousands separator. Otherwise treat as decimal.
            tail = s[last_comma + 1:]
            if len(tail) == 3 and tail.isdigit():
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")
        # only-dot or no-separator: leave as is

    try:
        v = float(s)
    except (TypeError, ValueError):
        return None
    return -v if is_neg else v


def _detect_weight_unit(values: list[float]) -> str:
    """Inspect a column of parsed weights and guess if they're percent or fraction.

    Returns ``"percent"`` if max ≤ 100 and sum > 5 (typical sum near 100),
    ``"fraction"`` if max ≤ 1.5, otherwise ``"percent"`` as the safer default.
    """
    if not values:
        return "percent"
    mx  = max(values)
    if mx <= 1.5:
        return "fraction"
    return "percent"


# ---------------------------------------------------------------------------
# Source resolution — URL or filesystem path
# ---------------------------------------------------------------------------
# The upload modal accepts a single "Source" string from the user. This
# can be:
#
#   * an http(s) URL                          → server fetches over the network
#   * a ``file://`` URI                       → server reads the path it points to
#   * a Windows-native absolute path          → server reads it directly
#   * a POSIX absolute path                   → server reads it directly
#
# Both branches return ``(filename, bytes)`` so the rest of the pipeline
# (parse → preview → mapping → commit) is source-agnostic. The original
# source string is preserved separately on the token under
# ``source_kind`` / ``source_value`` so we can write it back into the
# fund's upload_prefs after a successful commit.

# Cap on bytes we'll pull from a URL or open from disk. Holdings files
# are typically tens to hundreds of KB; even an iShares CSV with 7000
# bonds is well under 5 MB. The cap protects against accidental URLs
# that point to enormous payloads.
_SOURCE_BYTES_CAP = 25 * 1024 * 1024  # 25 MB

# Custom UA — some issuer CDNs return 403/404 for Python's default UA.
# This mimics a recent desktop browser and is enough for public CSVs.
_SOURCE_HTTP_UA = (
    "Mozilla/5.0 (PorxPy holdings importer; contact: localhost) "
    "AppleWebKit/537.36 Safari/537.36"
)


def _looks_like_url(s: str) -> bool:
    """Return True if ``s`` starts with an http(s) scheme."""
    return s.startswith(("http://", "https://"))


def _looks_like_file_uri(s: str) -> bool:
    """Return True if ``s`` is a ``file://`` URI."""
    return s.startswith("file://")


def _fileuri_to_path(uri: str) -> str:
    """Extract a filesystem path from a ``file://`` URI.

    Handles the common forms:

    * ``file:///C:/Users/me/x.csv``  → ``C:/Users/me/x.csv``  (Windows, three slashes)
    * ``file:///home/me/x.csv``      → ``/home/me/x.csv``     (POSIX, three slashes)
    * ``file://localhost/home/me/x`` → ``/home/me/x``         (rare, with host)

    Args:
        uri: A string starting with ``file://``.

    Returns:
        Path string suitable for :class:`pathlib.Path`. The result still
        needs to exist; the caller validates that.
    """
    rest = uri[len("file://"):]
    # Strip a leading "localhost" host segment if present.
    if rest.startswith("localhost/"):
        rest = rest[len("localhost"):]   # leaves "/home/..."
    # On Windows, file:///C:/... — drop the leading slash so "C:/..." is
    # what we open. On POSIX, file:///home/... — keep the leading slash.
    if len(rest) >= 3 and rest[0] == "/" and rest[2] == ":":
        rest = rest[1:]                  # "/C:/..." → "C:/..."
    # Decode percent-escapes (spaces, accented chars in paths)
    from urllib.parse import unquote
    return unquote(rest)


def _fetch_url_bytes(url: str) -> tuple[str, bytes]:
    """Download a URL and return ``(filename, bytes)``.

    Filename derivation:

    1. ``Content-Disposition: filename=...`` if the response sets one.
    2. Last segment of the URL path (after ``?`` is stripped).
    3. ``"download"`` as a final fallback.

    Raises:
        ValueError: HTTP error, redirect loop, payload too large, or empty.
    """
    import urllib.request
    import urllib.error
    from urllib.parse import urlparse, unquote

    req = urllib.request.Request(url, headers={
        "User-Agent": _SOURCE_HTTP_UA,
        "Accept":     "*/*",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            # Read with cap — bail if payload exceeds the limit so we
            # don't OOM on a huge response.
            data = resp.read(_SOURCE_BYTES_CAP + 1)
            if len(data) > _SOURCE_BYTES_CAP:
                raise ValueError(
                    f"download exceeds {_SOURCE_BYTES_CAP // (1024*1024)} MB cap"
                )
            # Filename from Content-Disposition (RFC 6266 simplified)
            cd = resp.headers.get("Content-Disposition") or ""
            fname = ""
            if "filename=" in cd.lower():
                # Quick parse — handles filename="x.csv" and filename=x.csv
                m = re.search(r'filename\*?=(?:UTF-\d\'\')?["]?([^";\r\n]+)', cd, re.IGNORECASE)
                if m:
                    fname = unquote(m.group(1)).strip().strip('"')
            if not fname:
                # Fall back to the URL path's basename
                pth = urlparse(resp.url or url).path
                fname = unquote(pth.rsplit("/", 1)[-1] or "download")
    except urllib.error.HTTPError as exc:
        raise ValueError(f"HTTP {exc.code} fetching {url}") from exc
    except urllib.error.URLError as exc:
        raise ValueError(f"network error fetching {url}: {exc.reason}") from exc

    if not data:
        raise ValueError(f"empty response from {url}")
    return fname, data


def _read_disk_bytes(path_str: str) -> tuple[str, bytes]:
    """Read a file from the local filesystem and return ``(filename, bytes)``.

    Single-user app, so no sandboxing — any readable path is accepted.

    Raises:
        ValueError: Path missing, not a regular file, too large, or unreadable.
    """
    p = Path(path_str).expanduser()
    if not p.exists():
        raise ValueError(f"path not found: {path_str}")
    if not p.is_file():
        raise ValueError(f"not a regular file: {path_str}")
    try:
        size = p.stat().st_size
    except OSError as exc:
        raise ValueError(f"cannot stat {path_str}: {exc}") from exc
    if size > _SOURCE_BYTES_CAP:
        raise ValueError(
            f"file size {size} exceeds {_SOURCE_BYTES_CAP // (1024*1024)} MB cap"
        )
    try:
        with open(p, "rb") as f:
            data = f.read()
    except OSError as exc:
        raise ValueError(f"cannot read {path_str}: {exc}") from exc
    if not data:
        raise ValueError(f"empty file: {path_str}")
    return p.name, data


def resolve_source(source: str) -> tuple[str, bytes, str]:
    """Resolve a free-form source string to bytes plus a kind tag.

    Args:
        source: User-provided source string. Whitespace is trimmed.

    Returns:
        ``(filename, raw_bytes, kind)`` where ``kind`` is ``"url"`` or
        ``"disk"``. ``filename`` is whatever the URL/path advertises; the
        caller passes it on to :func:`upload_preview` for format detection.

    Raises:
        ValueError: Empty input, unknown scheme, or fetch/read failure.
    """
    s = (source or "").strip()
    if not s:
        raise ValueError("source is empty")

    if _looks_like_url(s):
        fname, data = _fetch_url_bytes(s)
        return fname, data, "url"

    if _looks_like_file_uri(s):
        path = _fileuri_to_path(s)
        fname, data = _read_disk_bytes(path)
        return fname, data, "disk"

    # Otherwise treat as a filesystem path. Accepts both POSIX
    # (/home/me/x.csv) and Windows-native (C:\Users\me\x.csv) forms.
    # Path() handles both seamlessly.
    fname, data = _read_disk_bytes(s)
    return fname, data, "disk"


# ---------------------------------------------------------------------------
# Public API — preview
# ---------------------------------------------------------------------------
def upload_preview(filename: str, data: bytes, *,
                   sheet_name: str | None = None,
                   delimiter: str | None = None,
                   source_kind: str | None = None,
                   source_value: str | None = None) -> dict:
    """Parse an uploaded file and stash a preview token.

    Args:
        filename: Original filename — used only to detect format
            (extension-based) and surface in the response.
        data: Raw upload bytes.
        sheet_name: For XLSX, optionally select a non-default sheet.
        delimiter: For CSV, optionally override the auto-sniff.
        source_kind: Optional ``"url"`` / ``"disk"`` tag, recorded on the
            token so :func:`upload_commit` can write it into the fund's
            ``upload_prefs`` cache.
        source_value: Optional original source string (URL or path),
            paired with ``source_kind`` for the same purpose.

    Returns:
        A dict with::

            {
              "token":      "<uuid>",
              "filename":   "...",
              "format":     "csv" | "xlsx",
              "encoding":   "utf-8" or null,
              "delimiter":  "," or null,
              "sheets":     [...] or null,
              "picked_sheet": "..." or null,
              "rows":       [[str,...], ...],   # first ~200 rows
              "total_rows": <int>,
              "header_row": <int auto-detected>,
              "source_kind":  "url" | "disk" | null,
              "source_value": "..." | null,
            }
    """
    _reap_expired()

    ext = (filename.rsplit(".", 1)[-1] or "").lower()
    if ext in ("xlsx", "xlsm"):
        fmt = "xlsx"
        rows, info = _parse_xlsx(data, sheet_name=sheet_name)
    elif ext in ("csv", "tsv", "txt"):
        fmt = "csv"
        rows, info = _parse_csv(data, delimiter=delimiter)
    else:
        # Try CSV first as a more permissive fallback (most issuer files
        # are CSV with weird extensions).
        try:
            fmt = "csv"
            rows, info = _parse_csv(data, delimiter=delimiter)
        except Exception as exc:
            raise ValueError(f"unsupported file format: .{ext}") from exc

    # Guard against absurdly large uploads — only the first 200 rows of
    # each upload are surfaced to the dialog. The full grid is still
    # written to the token file so commit can read every row.
    PREVIEW_ROW_LIMIT = 200
    preview_rows = rows[:PREVIEW_ROW_LIMIT]
    header_row   = _detect_header_row(preview_rows)

    token = uuid.uuid4().hex
    payload = {
        "token":         token,
        "filename":      filename,
        "format":        fmt,
        "encoding":      info.get("encoding"),
        "delimiter":     info.get("delimiter"),
        "sheets":        info.get("sheets"),
        "picked_sheet":  info.get("picked_sheet"),
        "rows":          rows,                  # full grid persisted
        "total_rows":    len(rows),
        "header_row":    header_row,
        "source_kind":   source_kind,
        "source_value":  source_value,
        "created_at":    now_iso(),
    }
    _save_token(token, payload)

    # Frontend response carries only the preview slice
    return {
        **{k: v for k, v in payload.items() if k != "rows"},
        "rows":         preview_rows,
        "preview_rows": len(preview_rows),
    }


def upload_preview_from_source(source: str, *,
                               sheet_name: str | None = None,
                               delimiter: str | None = None) -> dict:
    """Resolve a free-form source string and run :func:`upload_preview`.

    The ``source`` may be an http(s) URL, a ``file://`` URI, or any
    absolute filesystem path (POSIX or Windows-native). The detection
    logic lives in :func:`resolve_source`.

    The resolved ``kind`` (``"url"`` or ``"disk"``) and the original
    ``source`` string are recorded on the preview token so that
    :func:`upload_commit` can persist them into the fund's
    ``upload_prefs`` cache for next-time prefill.

    Args:
        source: URL or filesystem path. See :func:`resolve_source`.
        sheet_name: For XLSX, optional sheet override.
        delimiter: For CSV, optional delimiter override.

    Returns:
        Same shape as :func:`upload_preview`, with the source fields
        populated.
    """
    filename, data, kind = resolve_source(source)
    return upload_preview(
        filename, data,
        sheet_name=sheet_name,
        delimiter=delimiter,
        source_kind=kind,
        source_value=source.strip(),
    )


# ---------------------------------------------------------------------------
# Public API — commit
# ---------------------------------------------------------------------------
def upload_commit(token: str, *,
                  ticker: str,
                  isin: str | None,
                  mapping: dict,
                  header_row: int,
                  decimal: str = "auto",
                  weight_unit: str = "auto",
                  sheet_name: str | None = None,
                  delimiter: str | None = None,
                  defaults: dict | None = None,
                  enrich_fields: list[str] | None = None) -> dict:
    """Public entry point — wraps the implementation in a try/finally so
    the cancel-token registry is always cleaned up, regardless of how
    the commit exits (success, ValueError, UploadCancelled, etc.).
    Without this, a cancelled-then-retried token would hit a stale
    flag on first iteration of the enrichment loop and raise
    immediately.
    """
    try:
        return _upload_commit_impl(
            token,
            ticker=ticker, isin=isin, mapping=mapping,
            header_row=header_row, decimal=decimal,
            weight_unit=weight_unit, sheet_name=sheet_name,
            delimiter=delimiter, defaults=defaults,
            enrich_fields=enrich_fields,
        )
    finally:
        clear_cancel(token)


def _upload_commit_impl(token: str, *,
                  ticker: str,
                  isin: str | None,
                  mapping: dict,
                  header_row: int,
                  decimal: str = "auto",
                  weight_unit: str = "auto",
                  sheet_name: str | None = None,
                  delimiter: str | None = None,
                  defaults: dict | None = None,
                  enrich_fields: list[str] | None = None) -> dict:
    """Apply a column mapping to a previewed upload and write the cache.

    Args:
        token: Token returned by :func:`upload_preview`.
        ticker: Yahoo ticker for this fund — used as the cache key.
        isin: Optional ISIN to record alongside the cache for traceability.
        mapping: ``{canonical_field: source_column_index_or_None}`` —
            ``name`` and ``weight`` are required. Optional fields may be
            ``None`` to mean "this file has no such column".
        header_row: Zero-based index of the header row (rows above are
            ignored, rows below are data).
        decimal: ``"dot"`` / ``"comma"`` / ``"auto"`` — passed to
            :func:`_parse_number`.
        weight_unit: ``"percent"`` / ``"fraction"`` / ``"auto"``. ``"auto"``
            uses :func:`_detect_weight_unit` after parsing all values.
        sheet_name: For XLSX, optionally override which sheet to read.
            Ignored if it matches the previewed sheet (no re-parse needed).
        delimiter: For CSV, optionally override the previewed delimiter
            (rare: would only be needed if the user changed it after preview).
        defaults: ``{field: default_value_string}`` — applied to every row
            for fields whose ``mapping`` is ``None``. Ignored for fields
            whose mapping picks a real column. Per-field semantics:

            * ``country``: run through :func:`country_to_mstar` like the
              mapped column, with the raw value retained as-is on miss.
            * ``currency``: uppercased.
            * ``sector`` / ``asset_class``: stored verbatim (asset_class
              is expected to be one of the canonical enum values; the
              UI restricts it).
            * any other key: silently ignored.

        enrich_fields: list of optional fields (subset of ``{"sector",
            "country", "currency", "asset_class"}``) for which the server
            should look up Yahoo per-symbol info for every row that has a
            ticker — overwriting the mapped/default value when Yahoo has
            something. Bounded to :data:`_ENRICH_ROW_CAP` rows per call;
            beyond that, the cap is honoured and the warnings dict notes
            how many rows were skipped.

    Returns:
        ``{rows_written, weight_sum_pct, warnings: {...}, _provider, ...}``.
    """
    payload = _load_token(token)
    if payload is None:
        raise ValueError("preview token expired or not found — re-upload the file")

    # Snapshot the user's originally-requested enrich choices before
    # Pass 3 narrows them to "unmapped + enrichable" only. We persist
    # the original choice into upload_prefs so a subsequent upload that
    # re-orders the column mapping doesn't silently drop an enrichment
    # the user previously asked for.
    enrich_requested_original = list(enrich_fields or [])

    # Re-parse only when the user changed sheet / delimiter mid-flow.
    rows = payload["rows"]
    fmt  = payload["format"]
    if fmt == "xlsx" and sheet_name and sheet_name != payload.get("picked_sheet"):
        # We don't have the original bytes any more; for now we treat
        # this as an error rather than re-uploading. Could be relaxed
        # by storing bytes alongside the token if it ever becomes a
        # common case.
        raise ValueError("sheet change after preview not supported — re-upload the file")
    if fmt == "csv" and delimiter and delimiter != payload.get("delimiter"):
        raise ValueError("delimiter change after preview not supported — re-upload the file")

    # Validate mapping
    if mapping.get("name") is None or mapping.get("weight") is None:
        raise ValueError("name and weight columns are required")

    # Required column-index validation (must point to a real column index)
    n_cols = max((len(r) for r in rows), default=0)
    for k, v in mapping.items():
        if v is None:
            continue
        if not isinstance(v, int) or v < 0 or v >= n_cols:
            raise ValueError(f"mapping[{k!r}]={v!r} is out of range (0..{n_cols-1})")

    if not isinstance(header_row, int) or header_row < 0 or header_row >= len(rows):
        raise ValueError(f"header_row {header_row} out of range")

    data_rows = rows[header_row + 1:]

    # First pass — collect raw weight strings so we can auto-detect units
    raw_weights: list[float] = []
    for r in data_rows:
        col = mapping["weight"]
        if col >= len(r):
            continue
        v = _parse_number(r[col], decimal=decimal)
        if v is not None:
            raw_weights.append(v)

    detected_unit = _detect_weight_unit(raw_weights)
    if weight_unit == "auto":
        unit = detected_unit
    elif weight_unit in ("percent", "fraction"):
        unit = weight_unit
    else:
        unit = detected_unit

    # Second pass — build the holdings rows from the file mapping ONLY.
    # Enrichment and default-fill happen in dedicated passes after this
    # so the precedence order (file > enrichment > default) stays
    # explicit and easy to reason about.
    out_rows:    list[dict] = []
    skipped_no_name   = 0
    skipped_no_weight = 0
    invalid_isins     = 0
    unmapped_countries: list[str] = []

    OPT_FIELDS = ("ticker", "isin", "sector", "country", "currency",
                  "asset_class", "sub_class",
                  # Bond metadata (v0.12.7) — read verbatim from the
                  # mapped column. coerce_holdings_row at the bottom of
                  # this function turns the numerics into floats and
                  # normalises the dates to DD/mmm/YYYY; we don't have
                  # to do anything special here besides shuttle the
                  # cell contents into the row.
                  "duration", "maturity", "coupon", "effective_date")

    isin_re = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")

    for r in data_rows:
        def cell(field: str) -> str:
            col = mapping.get(field)
            if col is None or col >= len(r):
                return ""
            return (r[col] or "").strip()

        name = cell("name")
        if not name:
            skipped_no_name += 1
            continue

        w_raw = cell("weight")
        w_val = _parse_number(w_raw, decimal=decimal)
        if w_val is None:
            skipped_no_weight += 1
            continue
        # Convert to canonical percent (5.34 means 5.34%)
        weight_pct = w_val * 100.0 if unit == "fraction" else w_val

        row_out: dict[str, Any] = {
            "name":        name,
            "weight_pct":  weight_pct,
        }

        # Optional fields — file-mapped values only at this stage
        for field in OPT_FIELDS:
            if mapping.get(field) is None:
                continue
            v = cell(field)
            if not v:
                continue
            if field == "isin":
                vu = v.upper().replace(" ", "")
                if isin_re.match(vu):
                    row_out["isin"] = vu
                else:
                    invalid_isins += 1
            elif field == "ticker":
                # Light cleanup only at upload time; the variant-probe
                # chokepoint in get_symbol_info_cached decides the
                # actual Yahoo form against Yahoo's response.
                row_out["ticker"] = clean_holding_ticker_input(v)
            elif field == "currency":
                # Try alias resolution first ("yen" → "JPY", "us dollar"
                # → "USD"); falls back to uppercased raw on no-match.
                resolved = resolve_currency(v)
                row_out["currency"] = resolved if resolved else v.upper().strip()
            elif field == "country":
                # Resolve to canonical mstar form; fall back to raw
                mstar = country_to_mstar(v)
                if mstar:
                    row_out["country"] = mstar
                else:
                    row_out["country"] = v       # keep raw so user can see
                    unmapped_countries.append(v)
            elif field == "sector":
                # Try alias resolution against the Morningstar 11-sector
                # taxonomy ("Information Technology" → "technology");
                # falls back to the raw value lower-cased on no-match
                # so issuer-specific spellings still display rather than
                # silently disappearing.
                resolved = resolve_sector(v)
                row_out["sector"] = resolved if resolved else v.strip().lower()
            elif field == "asset_class":
                # Normalise to the lowercase holding enum (equity / bond
                # / cash / other). normalize_holding_asset_class maps any
                # spelling; an unrecognised non-blank value becomes
                # "other" rather than being dropped.
                ac = normalize_holding_asset_class(v)
                if ac:
                    row_out["asset_class"] = ac
            elif field == "sub_class":
                # Try alias resolution against the canonical sub-class
                # taxonomy; falls back to lower-cased raw on no-match.
                resolved = resolve_sub_class(v)
                row_out["sub_class"] = resolved if resolved else v.strip().lower()
            else:
                row_out[field] = v

        # Currency derivation when no currency column was mapped: derive
        # from the file-mapped country (if present). Runs before the
        # enrichment/default passes — they may still overwrite in the
        # default-fill pass when neither file nor derivation supplied
        # a value, but this handles the common "country mapped, currency
        # not mapped" case for free.
        if "currency" not in row_out and "country" in row_out:
            derived = country_to_currency(row_out["country"])
            if derived:
                row_out["currency"] = derived
                row_out["currency_derived"] = True

        out_rows.append(row_out)

    # ──────────────────────────────────────────────────────────────────
    # Pass 3 — Yahoo enrichment for fields the user explicitly requested.
    #
    # Only fields that were UNMAPPED (mapping is None) AND listed in
    # ``enrich_fields`` are eligible. Bypassed for rows without a ticker
    # — Yahoo can't help us identify cash residuals or swap collateral.
    # Per-row override semantics: enrichment wins over a previously-set
    # default, but is itself a no-op if Yahoo has nothing for that field
    # on that ticker. Backed by the same per-symbol cache the top-10
    # enrichment uses, so repeat uploads of the same file are fast.
    # ──────────────────────────────────────────────────────────────────
    enrich_fields = list(enrich_fields or [])
    enrich_fields = [
        f for f in enrich_fields
        if f in _ENRICHABLE_FIELDS and mapping.get(f) is None
    ]
    enriched_counts: dict[str, int] = {f: 0 for f in enrich_fields}
    enrich_skipped_no_ticker = 0
    enrich_skipped_over_cap  = 0
    # Tickers Yahoo didn't recognise (cache returned no useful data).
    # Useful diagnostic — typically a sign that the variant probe
    # in get_symbol_info_cached couldn't find a Yahoo form that
    # matched, or the issuer used a local code we don't generate
    # candidates for. Surfaced as a count plus a sample of up to 20
    # tickers so the user can spot the pattern.
    enrich_unrecognised: list[str] = []

    if enrich_fields:
        # Lazy import to avoid pulling extractors → utils → ... at module
        # load time when upload.py is imported by the route registration.
        from porxpy.extractors import get_symbol_info_cached
        from porxpy.utils import alias_get

        # Cancel checkpoint #1: early exit if the user already hit
        # Cancel between commit-fire and reaching the loop. Cheap.
        if is_cancelled(token):
            raise UploadCancelled(token)

        for i, row in enumerate(out_rows):
            # Cancel checkpoint #2: top of every iteration. The per-row
            # work (a Yahoo network call on cold cache) dominates the
            # commit's total time, so this is where the cancel needs
            # to land. Worst-case latency between Cancel click and
            # actual stop is one in-flight network call.
            if is_cancelled(token):
                raise UploadCancelled(token)
            if i >= _ENRICH_ROW_CAP:
                enrich_skipped_over_cap += 1
                continue
            sym = (row.get("ticker") or "").strip()
            if not sym:
                enrich_skipped_no_ticker += 1
                continue
            try:
                info = get_symbol_info_cached(sym)
            except Exception as exc:
                print(f"[upload-enrich] {sym} error: {exc}")
                continue
            if not isinstance(info, dict):
                continue
            # Track Yahoo coverage. _found is set by get_symbol_info_cached;
            # falsy means Yahoo had no useful data for this ticker (and the
            # negative result has already been alias-cached so the next
            # upload won't re-probe).
            if not info.get("_found"):
                enrich_unrecognised.append(sym)
                continue   # nothing to enrich from

            # If the resolved ticker differs from what was on the row
            # (i.e. variant probing rewrote it — PLTRUS → PLTR), update
            # the row to the canonical Yahoo form. The holdings table
            # then shows clean tickers, and downstream rollup/lookups
            # see consistent symbols. The alias cache preserves the
            # original mapping for audit.
            present, resolved = alias_get(sym)
            if present and resolved and resolved != sym:
                row["ticker"] = resolved

            for f in enrich_fields:
                v = info.get(f)
                if not v:
                    continue
                if f == "country":
                    mstar = country_to_mstar(v) or v
                    row["country"] = mstar
                    enriched_counts["country"] += 1
                elif f == "currency":
                    row["currency"] = str(v).upper().strip()
                    enriched_counts["currency"] += 1
                else:
                    row[f] = v
                    enriched_counts[f] += 1

    # ──────────────────────────────────────────────────────────────────
    # Pass 4 — default-fill for unmapped fields the user supplied a
    # default value for. Runs LAST so it can't clobber file-mapped or
    # enriched data — only fills truly empty cells.
    # ──────────────────────────────────────────────────────────────────
    defaults = defaults or {}
    DEFAULTABLE = ("sector", "country", "currency", "asset_class", "sub_class",
                   # Bond metadata (v0.12.7) — defaultable in the
                   # "user typed a value in the upload-dialog default
                   # input" sense. No Yahoo enrichment exists for any
                   # of these (no _ENRICHABLE_FIELDS membership), but
                   # a fixed-coupon issuer-fund spreadsheet that omits
                   # the bond columns can still get them filled this
                   # way. coerce_holdings_row will type-coerce later.
                   "duration", "maturity", "coupon", "effective_date")
    default_apply: dict[str, str] = {}
    for f in DEFAULTABLE:
        if mapping.get(f) is not None:
            continue                                   # column was mapped
        raw = (defaults.get(f) or "").strip()
        if not raw:
            continue
        if f == "currency":
            resolved = resolve_currency(raw)
            default_apply[f] = resolved if resolved else raw.upper()
        elif f == "country":
            mstar = country_to_mstar(raw)
            default_apply[f] = mstar or raw
        elif f == "sector":
            resolved = resolve_sector(raw)
            default_apply[f] = resolved if resolved else raw.lower()
        elif f == "asset_class":
            # Default asset class is one of the holding enum values
            # (the upload dialog restricts the dropdown to them), but
            # normalise defensively in case a non-JSON caller sends
            # something else.
            ac = normalize_holding_asset_class(raw)
            if ac:
                default_apply[f] = ac
        elif f == "sub_class":
            resolved = resolve_sub_class(raw)
            default_apply[f] = resolved if resolved else raw.lower()
        else:
            # Includes the four bond fields — store the raw user value;
            # coerce_holdings_row normalises numerics and dates later.
            default_apply[f] = raw

    if default_apply:
        for row in out_rows:
            for f, val in default_apply.items():
                cur = row.get(f)
                # "Empty" means missing entirely OR a string that's
                # blank/whitespace. Any other truthy value (a number,
                # a non-blank string) is left alone — file mapping or
                # enrichment already populated it.
                is_empty = cur is None or (isinstance(cur, str) and not cur.strip())
                if is_empty:
                    row[f] = val
                    row.setdefault("_defaulted", []).append(f)

    # ──────────────────────────────────────────────────────────────────
    # Pass 5 — fund-asset-class fallback. Any row whose asset_class is
    # STILL blank after file-mapping, enrichment, and the user's default
    # falls back to the *fund's* asset class (equity → equity,
    # fixed_income → bond, cash → cash, else → other).
    #
    # The fund's asset class comes from its own ``asset_class`` cache
    # slot. If that slot is empty (the fund has never been loaded), there
    # is nothing to fall back to — the holding's asset_class stays blank,
    # and so does its sub_class (the two are a pair; coerce_holdings_row
    # only defaults sub_class when asset_class is set).
    #
    # sub_class itself is NOT defaulted here — coerce_holdings_row does
    # it uniformly for every row at the end, from whatever asset_class is
    # in effect by then (file / enrich / user default / this fallback).
    # ──────────────────────────────────────────────────────────────────
    fund_holding_ac = ""
    try:
        # Asset class is a fund-level category (ISIN-keyed) under the
        # 0.12.0 cache split.
        fund_blob = cache_read(isin, "asset_class") if isin else {}
        fund_ac_entry = (fund_blob.get("asset_class") or {}).get("value") or {}
        if isinstance(fund_ac_entry, dict):
            fund_holding_ac = default_holding_asset_class(
                fund_ac_entry.get("class"))
    except Exception as exc:
        print(f"[upload] fund asset-class lookup failed for {ticker}: {exc}")

    fund_ac_fallback_count = 0
    if fund_holding_ac:
        for row in out_rows:
            cur = row.get("asset_class")
            is_empty = cur is None or (isinstance(cur, str) and not cur.strip())
            if is_empty:
                row["asset_class"] = fund_holding_ac
                row.setdefault("_defaulted", []).append("asset_class")
                fund_ac_fallback_count += 1

    # Compute summary statistics
    weight_sum = sum(r.get("weight_pct", 0.0) for r in out_rows)

    # Normalise every row to the unified holdings superset schema and
    # stamp a stable ``_row_id`` on each (the holdings editor addresses
    # rows by this id). coerce_holdings_row preserves provenance extras
    # like ``_defaulted`` / ``currency_derived`` that the passes above
    # attached.
    out_rows = [coerce_holdings_row(r) for r in out_rows]

    # Build the unified ``holdings`` cache blob. ``source="manual_upload"``
    # marks this as the full user-provided list; the Yahoo fetch path in
    # load_fund_data will never overwrite a manual_upload blob (it's the
    # source of truth until the user re-uploads or clears it).
    holdings_blob = {
        "rows":        out_rows,
        "source":      "manual_upload",
        "_provider":   "manual",
        "row_count":   len(out_rows),
        "weight_sum_pct": round(weight_sum, 6),
        "isin_used":   isin or "",
        "uploaded_at": now_iso(),
        "filename":    payload.get("filename"),
        "weight_unit_used": unit,
        "decimal_used": decimal,
    }

    # Cancel checkpoint #3: last chance, just before the cache write.
    # Catches a cancel that landed during the final in-flight Yahoo
    # call or the post-loop coercion. After this point the holdings
    # are written; cancelling later is a no-op (the work is done).
    if is_cancelled(token):
        raise UploadCancelled(token)

    # Write to the funds cache (ISIN-keyed) under the unified
    # ``holdings`` slot. Holdings are fund-level — every listing of one
    # fund shares them — so this write makes the upload visible from
    # any of the fund's listings (BATT.L / BATG.L / etc).
    if not isin:
        raise ValueError("upload commit requires an ISIN (mode-2 funds "
                         "need their key supplied before holdings upload)")
    cache_put(isin, "holdings", holdings_blob)

    # ──────────────────────────────────────────────────────────────────
    # Persist the user's upload-dialog choices into the fund's
    # ``upload_prefs`` cache slot so the next "Upload holdings" for this
    # fund can prefill the source field, the column mapping, and all the
    # parsing knobs. Source kind/value were stashed on the token by the
    # /preview call; everything else came in on the commit body.
    #
    # Mapping is normalised to plain ints / nulls (the JSON layer does
    # this already, but commits coming from non-JSON callers might pass
    # numeric strings — be defensive).
    # ──────────────────────────────────────────────────────────────────
    norm_mapping: dict[str, Any] = {}
    for k, v in (mapping or {}).items():
        if v is None or v == "":
            norm_mapping[k] = None
        else:
            try:
                norm_mapping[k] = int(v)
            except (TypeError, ValueError):
                norm_mapping[k] = None

    prefs_blob = {
        "source_kind":   payload.get("source_kind"),
        "source_value":  payload.get("source_value"),
        "filename":      payload.get("filename"),
        "mapping":       norm_mapping,
        "header_row":    int(header_row) if isinstance(header_row, int) else 0,
        "decimal":       decimal,
        "weight_unit":   weight_unit,
        "defaults":      dict(defaults or {}),
        "enrich_fields": enrich_requested_original,
        "saved_at":      now_iso(),
    }
    cache_put(ticker, "upload_prefs", prefs_blob)

    # Cleanup
    _delete_token(token)

    # Unmatched-facet summary (v0.15.5). coerce_holdings_row stamps
    # row["_unmatched_facets"] with the field names that didn't resolve
    # against the resource CSVs (e.g. a bogus sector spelling). Up to
    # now this only surfaced in the portfolio-level review banner —
    # users uploading bogus values from a fund outside any portfolio
    # saw nothing. Roll the rows up here so the upload-dialog response
    # carries an explicit "X values need review" signal even before
    # the fund is added to a portfolio.
    uf_by_facet: dict[str, int] = {}
    uf_samples:  dict[str, list[str]] = {}
    for r in out_rows:
        for facet in (r.get("_unmatched_facets") or []):
            uf_by_facet[facet] = uf_by_facet.get(facet, 0) + 1
            raw = (r.get(facet) or "").strip()
            if raw:
                bucket = uf_samples.setdefault(facet, [])
                if raw not in bucket and len(bucket) < 5:
                    bucket.append(raw)
    uf_total = sum(uf_by_facet.values())

    # Per-row warnings — surfaced to the UI so the user knows what to fix
    warnings: dict[str, Any] = {
        "skipped_no_name":   skipped_no_name,
        "skipped_no_weight": skipped_no_weight,
        "invalid_isins":     invalid_isins,
        "unmapped_countries": sorted(set(unmapped_countries))[:50],
        # Cross-facet review summary (sector / sub_class / currency /
        # country). ``total`` is the count of (row, facet) pairs that
        # didn't resolve, not the count of distinct raw values — a
        # single misspelt sector across 30 rows counts as 30 here.
        "unmatched_facets": {
            "total":    uf_total,
            "by_facet": uf_by_facet,
            "samples":  uf_samples,
        },
        "enrichment": {
            "fields":              enrich_fields,
            "rows_filled":         enriched_counts,
            "skipped_no_ticker":   enrich_skipped_no_ticker,
            "skipped_over_cap":    enrich_skipped_over_cap,
            "row_cap":             _ENRICH_ROW_CAP,
            "unrecognised_count":  len(enrich_unrecognised),
            # First N as a sample so the user can spot a pattern
            # (e.g. all the unrecognised tickers come from one
            # exchange we don't normalise yet).
            "unrecognised_sample": sorted(set(enrich_unrecognised))[:20],
        },
        "defaults_applied":   sorted(default_apply.keys()),
        # Rows whose asset_class was blank after file/enrich/user-default
        # and fell back to the fund's asset class. 0 when the fund's
        # asset class wasn't known (nothing to fall back to) or every
        # row already had one.
        "asset_class_from_fund": fund_ac_fallback_count,
    }

    return {
        "rows_written":      len(out_rows),
        "weight_sum_pct":    round(weight_sum, 4),
        "weight_unit_used":  unit,
        "decimal_used":      decimal,
        "warnings":          warnings,
        "_provider":         "manual",
    }


def upload_clear(isin: str) -> bool:
    """Delete the manually-uploaded holdings for ``isin``.

    Used by the UI's "remove uploaded holdings" action. Holdings are
    fund-level (ISIN-keyed) under the 0.12.0 cache split, so this takes
    an ISIN — the caller (typically a route handler) resolves a ticker
    via :func:`listing_identity_lookup_isin` before calling.

    Only a ``manual_upload``-sourced blob is removed — if the slot
    currently holds Yahoo-sourced rows there's nothing for this action
    to do, and we leave them in place.

    Returns:
        ``True`` if a manual-upload holdings blob existed and was
        removed, ``False`` otherwise.
    """
    if not isin:
        return False
    blob = cache_read(isin, "holdings")
    entry = blob.get("holdings")
    if not entry:
        return False
    val = entry.get("value") or {}
    if not isinstance(val, dict) or val.get("source") != "manual_upload":
        return False
    del blob["holdings"]
    cache_write(isin, "holdings", blob)
    return True


# ---------------------------------------------------------------------------
# Helpers used by the cache-list endpoint and by extractors
# ---------------------------------------------------------------------------
def has_manual_holdings(isin: str) -> bool:
    """Return True iff the fund's holdings slot holds a manual upload.

    Holdings are fund-level (ISIN-keyed) under the 0.12.0 cache split.
    Used by the UI / portfolio loop to badge a fund with FULL · N when
    the slot was populated by a user upload rather than Yahoo.
    """
    if not isin:
        return False
    blob = cache_read(isin, "holdings")
    val = (blob.get("holdings") or {}).get("value") or {}
    return isinstance(val, dict) and val.get("source") == "manual_upload"


# ---------------------------------------------------------------------------
# Upload prefs — read / clear
# ---------------------------------------------------------------------------
def get_upload_prefs(ticker: str) -> dict | None:
    """Return the saved upload-dialog prefs for ``ticker``, or ``None``.

    Used by ``GET /api/upload/prefs`` so the upload modal can prefill
    the source field, mapping, defaults, etc. on a re-upload. Upload
    prefs are a per-LISTING UI memory (different listings of one fund
    may have different column layouts in their source spreadsheets),
    so they stay ticker-keyed under the 0.12.0 split.

    The TTL on ``upload_prefs`` is 3650 days (effectively forever) so
    a hit here is the same as "user has uploaded this fund before".
    """
    blob = cache_read(ticker, "upload_prefs")
    entry = blob.get("upload_prefs")
    if not entry:
        return None
    val = entry.get("value")
    if not isinstance(val, dict):
        return None
    return val


def clear_upload_prefs(ticker: str) -> bool:
    """Remove the saved upload-dialog prefs for ``ticker``.

    Currently only used implicitly via ``cache_purge`` and the UI's
    "Remove uploaded holdings" button doesn't touch this entry — the
    prefs are intentionally sticky so the next upload remembers the
    last-good mapping even if the user clears the holdings cache.
    Exposed here for parity with :func:`upload_clear` in case the UI
    grows a dedicated "forget upload settings" affordance.

    Returns:
        ``True`` if a prefs entry existed and was removed, ``False``
        if nothing was cached.
    """
    blob = cache_read(ticker, "upload_prefs")
    if "upload_prefs" not in blob:
        return False
    del blob["upload_prefs"]
    cache_write(ticker, "upload_prefs", blob)
    return True


# ---------------------------------------------------------------------------
# Breakdown-CSV upload pipeline (v0.16.0)
# ---------------------------------------------------------------------------
# Separate from the holdings pipeline above. Different shape, different
# storage (uploaded_breakdowns cache, not holdings), but reuses the same
# token-based preview/commit pattern.
#
# Flow:
#   1. Preview (parse_breakdown_csv_preview). Decode + sniff + parse the
#      CSV into ``[facet, key, weight]`` triples. Run each ``facet`` and
#      ``key`` through the canonical resolvers. Emit:
#        * ``accepted`` — the items that resolved cleanly, grouped by
#          canonical facet.
#        * ``unresolved_facets`` — distinct facet-column values that
#          didn't match {asset_class, sector, country, currency} after
#          case-insensitive whitespace-normalised comparison. The user
#          maps each to one of the four canonical facets in the
#          resolution modal.
#        * ``unresolved_keys`` — distinct (facet, key) pairs whose key
#          didn't canonicalise. Per facet the user picks a canonical
#          value from the dropdown (no per-row drop, no keep-as-is).
#      A token is stashed only if there's anything unresolved; otherwise
#      the caller goes straight to commit using the inline ``accepted``
#      payload.
#   2. Commit (commit_breakdown_upload). Take the token + the user's
#      resolution maps, apply them to the unresolved items, sum
#      duplicate (facet, key) weights after canonicalisation, normalise
#      percent → fraction per facet, and write to the
#      uploaded_breakdowns cache.
#
# The CSV schema is a long format with the columns:
#     facet, key, weight
# in any order (header row required). Extra columns are ignored.
# The weight column is parsed with the existing :func:`_parse_number`
# (auto decimal-notation) and the per-facet unit (percent vs fraction)
# is decided per-facet by :func:`_detect_weight_unit`. Zero is allowed
# (the row is silently omitted from the output); negative weights cause
# an outright rejection of the upload.

# Canonical facet names. The breakdown CSV's facet column must resolve
# to one of these. Case-insensitive whitespace-normalised lookup folds
# the obvious variants ("Asset Class", "asset_class", "asset-class")
# automatically; anything else surfaces in unresolved_facets.
_BD_FACETS: tuple[str, ...] = ("asset_class", "sector", "country", "currency")

# Built-in facet-name aliases that resolve without prompting the user.
# Keys are case-folded with non-alphanumerics stripped — so "Asset-Class",
# "asset class", "ASSET_CLASS" all collapse to the same lookup key. The
# set is intentionally narrow: only unambiguous case/spacing/punctuation
# variants of the four canonical facet names. Anything semantically
# different (e.g. "region", "industry", "geography") routes through the
# resolution modal so the user explicitly picks how it maps.
_BD_FACET_BUILTIN_ALIASES: dict[str, str] = {
    "assetclass":   "asset_class",
    "assetclasses": "asset_class",
    "sector":       "sector",
    "sectors":      "sector",
    "country":      "country",
    "countries":    "country",
    "currency":     "currency",
    "currencies":   "currency",
}


def _canon_facet_key(raw: Any) -> str:
    """Collapse a facet-column value to its alias-lookup form.

    Lowercases, strips non-alphanumerics. ``"Asset Class"`` → ``"assetclass"``.
    Used to match the user's facet column against
    :data:`_BD_FACET_BUILTIN_ALIASES`.
    """
    s = "" if raw is None else str(raw)
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _resolve_breakdown_key(facet: str, raw: str) -> str | None:
    """Resolve a raw key for ``facet`` to its canonical form.

    Args:
        facet: One of :data:`_BD_FACETS`.
        raw: The CSV's key cell (already stripped).

    Returns:
        Canonical key string when ``raw`` matches a known canonical or
        alias for ``facet``; ``None`` when it doesn't (caller routes to
        the resolution modal).

        * asset_class — :func:`porxpy.resources.resolve_asset_class`.
        * sector      — :func:`porxpy.resources.resolve_sector`.
        * country     — :func:`porxpy.resources.country_to_mstar`.
        * currency    — :func:`porxpy.resources.resolve_currency`.
    """
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    # Local imports so this module stays import-time-light and we can
    # be tolerant of resources reloads.
    from porxpy.resources import (
        country_to_mstar, resolve_asset_class, resolve_currency,
        resolve_sector,
    )
    if facet == "asset_class":
        return resolve_asset_class(raw)
    if facet == "sector":
        return resolve_sector(raw)
    if facet == "country":
        return country_to_mstar(raw)
    if facet == "currency":
        return resolve_currency(raw)
    return None


def parse_breakdown_csv_preview(
    filename: str,
    data: bytes,
    delimiter: str | None = None,
) -> dict:
    """Parse a breakdown CSV and identify what (if anything) needs user mapping.

    Args:
        filename: Original upload filename (used for error messages).
        data: Raw upload bytes.
        delimiter: Optional CSV delimiter override; otherwise auto-sniff.

    Returns:
        ::

            {
              "token":              str | None,
              "filename":           str,
              "encoding":           str,
              "delimiter":          str,
              "row_count":          int,
              "accepted": {
                  "asset_class": [{"key","weight"}, ...],
                  "sector":      [...], "country": [...], "currency": [...],
              },
              "unresolved_facets":  [{"raw": "...", "rows": int}, ...],
              "unresolved_keys": [
                  {"facet": "country", "raw": "...", "rows": int}, ...
              ],
              "warnings":           [str, ...],
            }

        ``token`` is None when the file resolved cleanly with no user
        action needed (the caller can commit immediately using
        ``accepted``). When unresolved items exist, the token must be
        passed back to :func:`commit_breakdown_upload` alongside the
        user's facet-mapping and key-mapping decisions.

    Raises:
        ValueError: On structural problems — empty file, header missing,
            required column missing, negative weights, fully unparseable
            numeric column, etc. Caller surfaces the message to the user.
    """
    if not data:
        raise ValueError("the uploaded file is empty")

    # Parse CSV ---------------------------------------------------------
    rows, info = _parse_csv(data, delimiter=delimiter)
    # Strip blank rows and trim cells.
    grid: list[list[str]] = []
    for r in rows:
        cleaned = [("" if c is None else str(c)).strip() for c in r]
        if any(cleaned):
            grid.append(cleaned)
    if not grid:
        raise ValueError("the uploaded file has no non-empty rows")

    # Header detection: the first non-empty row is the header. We don't
    # try anything fancier — the LLM-screenshot use case produces clean
    # 3-column files. The user can always fix the file if they put a
    # preamble in.
    header_raw = grid[0]
    header = [c.lower() for c in header_raw]

    def _find_col(*candidates: str) -> int | None:
        # Case-insensitive, dash/underscore/whitespace-collapsed.
        wanted = {re.sub(r"[^a-z0-9]", "", w.lower()) for w in candidates}
        for i, h in enumerate(header):
            if re.sub(r"[^a-z0-9]", "", h) in wanted:
                return i
        return None

    facet_col  = _find_col("facet", "category", "type")
    key_col    = _find_col("key", "name", "label", "value")
    weight_col = _find_col("weight", "percent", "percentage", "pct",
                           "share", "allocation")

    missing = []
    if facet_col is None:
        missing.append("facet")
    if key_col is None:
        missing.append("key")
    if weight_col is None:
        missing.append("weight")
    if missing:
        raise ValueError(
            f"the CSV is missing required column(s): {', '.join(missing)}. "
            f"Header row read as: {header_raw!r}. The CSV needs columns "
            f"named 'facet', 'key', and 'weight' (any order).")

    # Parse rows --------------------------------------------------------
    # Track per-facet items (raw form), unresolved facet-column values,
    # and unresolved (facet, key) pairs. Counts on the unresolved
    # buckets help the frontend say "this affects N rows".
    data_rows = grid[1:]
    if not data_rows:
        raise ValueError("the CSV has a header but no data rows")

    # Pass 1: collect raw rows.
    parsed: list[tuple[str, str, float]] = []  # (facet_raw, key_raw, weight)
    bad_weight_rows: list[int] = []
    negative_rows: list[int] = []
    for i, r in enumerate(data_rows, start=2):  # +2 because of header + 1-based
        if facet_col >= len(r) or key_col >= len(r) or weight_col >= len(r):
            continue  # short row; skip silently
        facet_raw = r[facet_col]
        key_raw   = r[key_col]
        w_raw     = r[weight_col]
        if not facet_raw and not key_raw and not w_raw:
            continue
        w = _parse_number(w_raw, decimal="auto")
        if w is None:
            bad_weight_rows.append(i)
            continue
        if w < 0:
            negative_rows.append(i)
            continue
        parsed.append((facet_raw, key_raw, float(w)))

    if negative_rows:
        sample = negative_rows[:5]
        raise ValueError(
            f"the CSV contains negative weight(s) on row "
            f"{'s ' if len(sample) > 1 else ''}{sample}. Negative "
            f"weights are not allowed; please fix the file and re-upload.")

    if not parsed:
        raise ValueError("the CSV has no parseable data rows")

    warnings: list[str] = []
    if bad_weight_rows:
        warnings.append(
            f"skipped {len(bad_weight_rows)} row(s) with unparseable "
            f"weight values (e.g. row {bad_weight_rows[0]})")

    # Pass 2: resolve facet column ------------------------------------
    # Group rows by canonical facet (or 'unresolved' bucket). Builtin
    # aliases resolve silently; anything else surfaces in
    # unresolved_facets for user mapping.
    facet_groups: dict[str, list[tuple[str, float]]] = {
        f: [] for f in _BD_FACETS
    }
    unresolved_facet_bucket: dict[str, list[tuple[str, float]]] = {}
    for facet_raw, key_raw, w in parsed:
        canon = _BD_FACET_BUILTIN_ALIASES.get(_canon_facet_key(facet_raw))
        if canon:
            facet_groups[canon].append((key_raw, w))
        else:
            unresolved_facet_bucket.setdefault(facet_raw, []).append((key_raw, w))

    unresolved_facets: list[dict] = [
        {"raw": raw, "rows": len(rows)}
        for raw, rows in sorted(
            unresolved_facet_bucket.items(), key=lambda kv: -len(kv[1]))
    ]

    # Pass 3: resolve keys within each canonical facet ----------------
    # accepted[facet]: list of (canonical_key, weight)
    # unresolved_keys: distinct (facet, raw_key) pairs awaiting mapping.
    accepted: dict[str, list[tuple[str, float]]] = {f: [] for f in _BD_FACETS}
    unresolved_keys_bucket: dict[tuple[str, str], int] = {}
    for facet, rows in facet_groups.items():
        for key_raw, w in rows:
            canon_key = _resolve_breakdown_key(facet, key_raw)
            if canon_key:
                accepted[facet].append((canon_key, w))
            else:
                k_raw = key_raw.strip()
                if not k_raw:
                    # Blank key with a non-blank weight — skip silently,
                    # nothing meaningful to do with it.
                    continue
                unresolved_keys_bucket[(facet, k_raw)] = (
                    unresolved_keys_bucket.get((facet, k_raw), 0) + 1)

    unresolved_keys: list[dict] = [
        {"facet": facet, "raw": raw, "rows": cnt}
        for (facet, raw), cnt in sorted(
            unresolved_keys_bucket.items(), key=lambda kv: (kv[0][0], -kv[1]))
    ]

    # Build accepted payload in the JSON-friendly shape. Note: we do NOT
    # collapse duplicate canonical keys here — that's done at commit
    # time (after the user's resolutions are merged in), so duplicates
    # introduced by user mappings get summed too.
    accepted_out: dict[str, list[dict]] = {
        f: [{"key": k, "weight": w} for k, w in items]
        for f, items in accepted.items()
    }

    needs_token = bool(unresolved_facets) or bool(unresolved_keys)
    token: str | None = None
    if needs_token:
        # Reap expired tokens first (cheap; same pattern as the
        # holdings preview).
        _reap_expired()
        token = uuid.uuid4().hex

        # Bucket unresolved keys by (facet, raw_key) → list of weights.
        # Multiple CSV rows may share a (facet, raw_key) pair (e.g. two
        # rows both spelling country "USA"); we record every weight so
        # commit sums them under the user's chosen canonical mapping.
        # Walk facet_groups, not the bucket dict, since the bucket only
        # carried row counts.
        unresolved_keys_payload: dict[str, dict[str, list[float]]] = {
            f: {} for f in _BD_FACETS
        }
        for facet, rows in facet_groups.items():
            for key_raw, w in rows:
                k_stripped = key_raw.strip()
                if not k_stripped:
                    continue
                if _resolve_breakdown_key(facet, k_stripped):
                    continue   # already accepted
                unresolved_keys_payload[facet].setdefault(
                    k_stripped, []).append(w)

        payload = {
            "kind":       "breakdown_upload",
            "filename":   filename,
            "created_at": now_iso(),
            "encoding":   info.get("encoding"),
            "delimiter":  info.get("delimiter"),
            # The items that already cleaned up — committed as-is.
            "accepted":   accepted_out,
            # Rows whose facet column needs user mapping. Stored
            # verbatim (raw key, weight); commit resolves keys after
            # the user picks the canonical facet.
            "unresolved_facets": {
                raw: [{"key": k, "weight": w} for k, w in rows]
                for raw, rows in unresolved_facet_bucket.items()
            },
            # Rows whose facet resolved but whose key didn't —
            # bucketed by canonical facet, then by raw key.
            "unresolved_keys": unresolved_keys_payload,
        }
        _save_token(token, payload)

    return {
        "token":             token,
        "filename":          filename,
        "encoding":          info.get("encoding"),
        "delimiter":         info.get("delimiter"),
        "row_count":         len(parsed),
        "accepted":          accepted_out,
        "unresolved_facets": unresolved_facets,
        "unresolved_keys":   unresolved_keys,
        "warnings":          warnings,
    }


def commit_breakdown_upload(
    isin: str,
    token: str | None,
    inline_accepted: dict | None,
    facet_map: dict | None,
    key_map: dict | None,
) -> dict:
    """Commit a breakdown CSV upload to the ``uploaded_breakdowns`` cache.

    Two call modes:

    * **No token** (``token is None``). The preview returned no
      unresolved items, so the frontend can commit immediately using
      the inline ``accepted`` payload. ``facet_map`` and ``key_map``
      are ignored.
    * **With token**. The preview stashed a payload because something
      needed mapping. The frontend gathered the user's choices and
      sends them as ``facet_map`` (raw-facet → canonical-facet) and
      ``key_map`` (``{facet: {raw_key: canonical_key}}``). Both maps
      are applied to the stashed payload before writing.

    Args:
        isin: Fund ISIN.
        token: Token returned by ``parse_breakdown_csv_preview``, or
            None to use ``inline_accepted``.
        inline_accepted: Per-facet accepted items (preview's output)
            when no token is needed.
        facet_map: ``{raw_facet: canonical_facet}``. Every raw value in
            the preview's ``unresolved_facets`` MUST be in this map and
            mapped to one of :data:`_BD_FACETS`.
        key_map: ``{facet: {raw_key: canonical_key}}``. Every
            ``(facet, raw)`` in the preview's ``unresolved_keys`` MUST
            be mapped to a canonical value for that facet.

    Returns:
        ::

            {
              "isin":     "...",
              "facets":   {facet: [{"key","weight"}, ...]},
              "weights":  {facet: "percent" | "fraction" | "empty"},
              "summary":  {facet: {"items": N, "weight_sum": 0.97}, ...},
            }

    Raises:
        ValueError: If the resolution maps don't cover every unresolved
            item, or if a mapped target isn't one of the canonical
            values.
    """
    # Local import to avoid a cycle at module-import time.
    from porxpy.utils import uploaded_breakdowns_put

    isin = (isin or "").strip().upper()
    if not isin:
        raise ValueError("isin is required")

    # Per-facet accumulator: list of (canonical_key, weight). Duplicates
    # are summed at the end (per the design — uploading "US" and
    # "United States" both → united_states means the user meant them
    # combined).
    per_facet: dict[str, list[tuple[str, float]]] = {f: [] for f in _BD_FACETS}

    # Branch 1: inline accepted (no resolution needed).
    if not token:
        if not isinstance(inline_accepted, dict):
            raise ValueError(
                "no token provided and no inline accepted payload")
        for facet in _BD_FACETS:
            for it in (inline_accepted.get(facet) or []):
                k = str(it.get("key") or "").strip()
                if not k:
                    continue
                try:
                    w = float(it.get("weight") or 0.0)
                except (TypeError, ValueError):
                    continue
                if w < 0:
                    raise ValueError(
                        "negative weight in inline accepted payload")
                per_facet[facet].append((k, w))

    # Branch 2: token-based resolution.
    else:
        payload = _load_token(token)
        if not payload or payload.get("kind") != "breakdown_upload":
            raise ValueError(
                "upload token has expired or is invalid; please re-upload")

        facet_map = facet_map or {}
        key_map   = key_map   or {}

        # Apply facet_map to unresolved facet rows -------------------
        # We require every key in unresolved_facets to appear in
        # facet_map and to map to one of _BD_FACETS.
        unresolved_facet_payload = payload.get("unresolved_facets") or {}
        missing_facets: list[str] = []
        bad_facet_targets: list[tuple[str, str]] = []
        # Bucket the still-raw rows by what the user mapped them to.
        facet_remapped: dict[str, list[tuple[str, float]]] = {
            f: [] for f in _BD_FACETS
        }
        for raw, rows in unresolved_facet_payload.items():
            target = facet_map.get(raw)
            if not target:
                missing_facets.append(raw)
                continue
            if target not in _BD_FACETS:
                bad_facet_targets.append((raw, target))
                continue
            for it in rows:
                facet_remapped[target].append(
                    (str(it.get("key") or ""), float(it.get("weight") or 0.0))
                )
        if missing_facets:
            raise ValueError(
                f"unresolved facet column value(s) not mapped: "
                f"{missing_facets!r}")
        if bad_facet_targets:
            raise ValueError(
                f"facet_map targets must be one of {list(_BD_FACETS)}; "
                f"got {bad_facet_targets!r}")

        # Start from the already-accepted items.
        for facet in _BD_FACETS:
            for it in (payload.get("accepted") or {}).get(facet, []):
                per_facet[facet].append(
                    (str(it.get("key") or ""), float(it.get("weight") or 0.0))
                )

        # Apply key_map within each facet ----------------------------
        # Re-resolve the keys on the freshly-bucketed rows (the ones we
        # just mapped from unresolved facets), AND map the originally-
        # unresolved keys per facet. Anything still without a canonical
        # form after user mapping is rejected.
        from porxpy.resources import (
            country_to_mstar, resolve_asset_class, resolve_currency,
            resolve_sector,
        )
        _per_facet_resolver = {
            "asset_class": resolve_asset_class,
            "sector":      resolve_sector,
            "country":     country_to_mstar,
            "currency":    resolve_currency,
        }

        unresolved_keys_payload = payload.get("unresolved_keys") or {}
        missing_keys: list[tuple[str, str]] = []
        bad_key_targets: list[tuple[str, str, str]] = []

        for facet in _BD_FACETS:
            resolver = _per_facet_resolver[facet]
            facet_key_map = (key_map.get(facet) or {})
            # 3a. The rows that came from a user-mapped facet column —
            # their keys haven't been resolver-checked yet.
            for k_raw, w in facet_remapped[facet]:
                k = k_raw.strip()
                if not k:
                    continue
                canon = resolver(k)
                if not canon:
                    canon = (facet_key_map.get(k) or "").strip()
                    if not canon:
                        missing_keys.append((facet, k))
                        continue
                    # Validate the user's target against the resolver.
                    # We allow the user to type a canonical form
                    # directly (e.g. "unitedstates") — the resolver will
                    # accept its own output. If it doesn't accept it,
                    # reject.
                    if not resolver(canon):
                        # Also accept literal canonical values (some
                        # resolvers don't accept their own output as
                        # input — country_to_mstar accepts both forms,
                        # but asset-class resolve uses HOLDINGS_CLASS).
                        if canon not in _allowed_canonical_set(facet):
                            bad_key_targets.append((facet, k, canon))
                            continue
                per_facet[facet].append((canon, float(w)))

            # 3b. The originally-unresolved keys for this facet.
            unresolved_for_facet = unresolved_keys_payload.get(facet) or {}
            for raw, weights in unresolved_for_facet.items():
                canon = (facet_key_map.get(raw) or "").strip()
                if not canon:
                    missing_keys.append((facet, raw))
                    continue
                if not resolver(canon):
                    if canon not in _allowed_canonical_set(facet):
                        bad_key_targets.append((facet, raw, canon))
                        continue
                for w in weights:
                    per_facet[facet].append((canon, float(w)))

        if missing_keys:
            sample = missing_keys[:5]
            raise ValueError(
                f"unresolved key(s) not mapped: "
                f"{sample}{' ...' if len(missing_keys) > 5 else ''}")
        if bad_key_targets:
            sample = bad_key_targets[:5]
            raise ValueError(
                f"key_map target(s) not a canonical value: "
                f"{sample}{' ...' if len(bad_key_targets) > 5 else ''}")

        # Token consumed.
        _delete_token(token)

    # Collapse duplicates and normalise weights -----------------------
    # Per facet: sum weights by canonical key (duplicates are summed,
    # per the user's design choice). Auto-detect percent vs fraction
    # per-facet from the summed values; if percent, divide by 100. Drop
    # items with zero weight (the design treats zero as "omit").
    facets_out: dict[str, list[dict]] = {f: [] for f in _BD_FACETS}
    weights_meta: dict[str, str] = {}
    summary: dict[str, dict] = {}

    for facet in _BD_FACETS:
        merged: dict[str, float] = {}
        for k, w in per_facet[facet]:
            merged[k] = merged.get(k, 0.0) + w
        # Drop zero-weight items.
        merged = {k: v for k, v in merged.items() if v > 0}
        if not merged:
            weights_meta[facet] = "empty"
            summary[facet] = {"items": 0, "weight_sum": 0.0}
            continue
        unit = _detect_weight_unit(list(merged.values()))
        weights_meta[facet] = unit
        divisor = 100.0 if unit == "percent" else 1.0
        items = [
            {"key": k, "weight": round(v / divisor, 6)}
            for k, v in sorted(merged.items(), key=lambda kv: -kv[1])
        ]
        facets_out[facet] = items
        summary[facet] = {
            "items":      len(items),
            "weight_sum": round(sum(it["weight"] for it in items), 6),
        }

    persisted = uploaded_breakdowns_put(isin, facets_out)

    return {
        "isin":    isin,
        "facets":  persisted,
        "weights": weights_meta,
        "summary": summary,
    }


def _allowed_canonical_set(facet: str) -> set[str]:
    """Return the set of canonical key values currently allowed for ``facet``.

    Used at commit-time to validate that a user's mapping target is a
    legitimate canonical value, even if it round-trips through the
    resolver oddly (some resolvers don't accept their own output as
    input — e.g. asset-class resolve uses the HOLDINGS_CLASS aliases).
    """
    from porxpy.resources import (
        CURRENCY_ROWS, HOLDINGS_CLASS_ROWS, SECTORS_ROWS,
        COUNTRY_ROWS,
    )
    if facet == "asset_class":
        out = set()
        for r in HOLDINGS_CLASS_ROWS:
            ac = (r.get("asset_class") or "").strip().lower()
            if ac:
                out.add(ac)
        return out
    if facet == "sector":
        return {(r.get("sector") or "").strip().lower()
                for r in SECTORS_ROWS if r.get("sector")}
    if facet == "country":
        return {(r.get("mstar_country") or "").strip().lower()
                for r in COUNTRY_ROWS if r.get("mstar_country")}
    if facet == "currency":
        return {(r.get("code") or "").strip()
                for r in CURRENCY_ROWS if r.get("code")}
    return set()


def list_canonical_values(facet: str) -> list[dict]:
    """Return the canonical value list for ``facet``, for dropdown population.

    Used by the resolution modal's frontend to drive the per-facet
    "map to" dropdowns.

    Args:
        facet: One of :data:`_BD_FACETS`.

    Returns:
        ``[{"key": canonical, "label": display_name}, ...]`` sorted by
        label. Empty list for unknown facets. Labels are intentionally
        plain strings (no HTML); the frontend's existing per-facet
        formatters (``fmtCountry`` / ``fmtSector`` / ``fmtAssetClass``)
        can be applied at render time if a richer label is wanted.
    """
    from porxpy.resources import (
        CURRENCY_ROWS, HOLDINGS_CLASS_ROWS, SECTORS_ROWS,
        COUNTRY_ROWS,
    )
    if facet == "asset_class":
        seen: set[str] = set()
        out: list[dict] = []
        for r in HOLDINGS_CLASS_ROWS:
            ac = (r.get("asset_class") or "").strip().lower()
            if ac and ac not in seen:
                seen.add(ac)
                out.append({"key": ac, "label": ac.replace("_", " ").title()})
        out.sort(key=lambda x: x["label"])
        return out
    if facet == "sector":
        out = []
        for r in SECTORS_ROWS:
            sec = (r.get("sector") or "").strip().lower()
            if sec:
                # SECTORS_ROWS carries no display_name — title-case the
                # canonical key for a human-friendly label.
                out.append({"key": sec, "label": sec.replace("_", " ").title()})
        out.sort(key=lambda x: x["label"])
        return out
    if facet == "country":
        out = []
        for r in COUNTRY_ROWS:
            ms = (r.get("mstar_country") or "").strip().lower()
            if ms:
                # mstar_country values are already a reasonable label
                # ("unitedstates" → "United States"). The frontend can
                # apply fmtCountry for nicer rendering; here we expose
                # a fallback title-cased form.
                out.append({"key": ms, "label": ms.title()})
        out.sort(key=lambda x: x["label"])
        return out
    if facet == "currency":
        out = []
        for r in CURRENCY_ROWS:
            code = (r.get("code") or "").strip()
            if code:
                name  = (r.get("name") or "").strip()
                label = f"{code} — {name}" if name else code
                out.append({"key": code, "label": label})
        out.sort(key=lambda x: x["label"])
        return out
    return []
